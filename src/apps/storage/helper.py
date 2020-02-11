import logging

import pyexcel

from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook

from src.apps.storage.models import Invoice, Product, Revise, ProductRevise
from src.apps.storage.service import get_or_create_group, get_or_create_category, get_or_create_kind, \
    update_or_create_product
from src.base_components.exceptions import ParseFileException
from src.template_tags.common_tags import format_date

logger = logging.getLogger('common_log')


class ExcelFileProcessor(object):

    def __init__(self, excel_file):
        self.excel_file = excel_file
        self.errors = []
        self.file_types = ['xls', 'xlsx']
        self.column_in_row = 0
        self.result = None
        self.start_log_message = 'empty start log message!'
        self.end_log_message = 'empty end log message!'

    def process(self):
        file_type = self.excel_file.name.split(".")[1]
        if file_type not in self.file_types:
            self.errors.append('Файл недопустимого формата, допустимые форматы - ' + str(self.file_types))
            return

        sheet = pyexcel.get_sheet(file_type=file_type, file_stream=self.excel_file.read())

        logger.info('***********************************************************')
        logger.info('Начинаем обработку файла %s %s' % (self.excel_file.name, self.start_log_message))
        logger.info('***********************************************************')
        self.process_business_logic(sheet.rows())
        logger.info('*******************************************************')
        logger.info('Обработка файла %s %s' % (self.excel_file.name, self.end_log_message))
        logger.info('*******************************************************')

    @transaction.atomic()
    def process_business_logic(self, rows):
        # пропускаем шапку
        head = True
        index = 1
        for row in rows:
            if not head:
                try:
                    index += 1
                    self.process_excel_row(row)
                except Exception as e:
                    message = 'Ошибка при обработке строки файла номер ' + str(index) + ' данные строки - ' + str(
                        row) + ' ошибка - ' + str(e)
                    logger.error(message)
                    self.errors.append(message)
            else:
                head = False

    def process_excel_row(self, row, **kwargs):
        logger.info("Начинаем обработку сырой строки - " + str(row))
        row = self.pre_process_row(row)
        logger.info("Строка после обработки - " + str(row))
        result = self.process_business_logic_for_row(row, **kwargs)
        logger.info("Строка успешно обработана!")
        return result

    def pre_process_row(self, row):
        if len(row) != self.column_in_row:
            raise ParseFileException(message=str(row) + ' - ' + u'в строке должно быть %s столбцов' % self.column_in_row)
        return ['0' if not item else item for item in row]  # пустые значения приравниваем к 0

    def process_business_logic_for_row(self, row, **kwargs):
        raise NotImplementedError()

    def get_errors(self):
        return self.errors

    def get_result(self):
        return self.result


class ProductExcelProcessor(ExcelFileProcessor):

    def __init__(self, excel_file):
        super(ProductExcelProcessor, self).__init__(excel_file)
        self.column_in_row = 11
        self.start_log_message = 'для загрузки остатков на склад!'
        self.end_log_message = 'остатков на склад завершена!'

    def process_business_logic_for_row(self, row, **kwargs):
        group = get_or_create_group(row[0])
        category = get_or_create_category(row[1], group)
        kind = get_or_create_kind(row[2], category)
        update_or_create_product(kind, row[3:])


class ReviseProductExcelProcessor(ExcelFileProcessor):

    def __init__(self, user, excel_file):
        super(ReviseProductExcelProcessor, self).__init__(excel_file)
        self.column_in_row = 7
        self.start_log_message = 'для загрузки сверки товара!'
        self.end_log_message = 'сверки товара завершена!'
        self.user = user

    @transaction.atomic()
    def process_business_logic(self, rows):
        if not rows:
            return

        revise = Revise(
            owner=self.user,
            revise_date=timezone.now()
        )
        revise.save()
        products_revise = []

        # пропускаем шапку
        head = True
        index = 1
        for row in rows:
            if not head:
                try:
                    index += 1
                    products_revise.append(self.process_excel_row(row, **{'revise_id': revise.id}))
                except Exception as e:
                    message = 'Ошибка при обработке строки файла номер ' + str(index) + ' данные строки - ' + str(
                        row) + ' ошибка - ' + str(e)
                    logger.error(message)
                    self.errors.append(message)
            else:
                head = False

        if self.errors:
            return

        ProductRevise.objects.bulk_create(products_revise)
        revise.calculate_loss()
        self.result = revise

    def process_business_logic_for_row(self, row, **kwargs):
        product = Product.objects.get(id=int(row[0]))
        count_revise = int(row[6])

        return ProductRevise(
            product=product,
            count_revise=count_revise if count_revise > 0 else 0,
            count_storage=product.product_count,
            revise_id=kwargs['revise_id']
        )


class ExportProductProcessor(object):
    def __init__(self, kinds=(), export_type=''):
        super(ExportProductProcessor, self).__init__()
        self.kinds = kinds
        self.export_type = export_type

    @staticmethod
    def post_process_sheet(sheet):
        dims = {}
        for row in sheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            sheet.column_dimensions[col].width = value + 5

    def __generate_for_restore(self):

        products = Product.objects.select_related().all().order_by('product_kind__product_category__product_group_id',
                                                                   'product_kind__product_category__category_name',
                                                                   'product_kind__kind_name', 'product_name')
        book = Workbook()
        sheet = book.create_sheet(title='Полная выгрузка товара', index=0)
        sheet.append(['Группа', 'Категория', 'Вид', 'Наименование', 'Закуп', 'Розница', 'Дисконт', 'Заведение', 'Оптом',
                      'Остаток', 'Мин.Кол'])

        for product in products:
            kind = product.product_kind
            category = kind.product_category

            row = [category.product_group.group_name, category.category_name, kind.kind_name,
                   product.product_name, product.cost_price, product.price_retail, product.price_discount,
                   product.price_shop, product.price_wholesale, product.product_count, product.min_count]
            sheet.append(row)
        self.post_process_sheet(sheet)
        return book

    def __generate_for_wholesales(self):
        products = Product.objects.select_related().filter(product_kind__in=self.kinds).filter(product_count__gt=0) \
            .order_by('product_kind__product_category__product_group_id',
                      'product_kind__product_category__category_name',
                      'product_kind__kind_name', 'product_name')
        book = Workbook()
        sheet = book.create_sheet(title='Остаток товаров для оптовиков', index=0)
        sheet.append(['Группа', 'Категория', 'Вид', 'Наименование', 'Остаток'])

        for product in products:
            kind = product.product_kind
            category = kind.product_category
            row = [category.product_group.group_name, category.category_name, kind.kind_name,
                   product.product_name, product.product_count]
            sheet.append(row)
        self.post_process_sheet(sheet)
        return book

    def __generate_for_revise(self):

        products = Product.objects.select_related().filter(product_kind__in=self.kinds) \
            .order_by('product_kind__product_category__product_group_id',
                      'product_kind__product_category__category_name',
                      'product_kind__kind_name', 'product_name')
        book = Workbook()
        sheet = book.create_sheet(title='Товар для сверки', index=0)
        sheet.append(['id', 'Группа', 'Категория', 'Вид', 'Наименование', 'Остаток в системе', 'Фактический остаток'])

        for product in products:
            kind = product.product_kind
            category = kind.product_category
            row = [product.id, category.product_group.group_name, category.category_name, kind.kind_name,
                   product.product_name, product.product_count, product.product_count]
            sheet.append(row)
        self.post_process_sheet(sheet)
        return book

    def generate_storage_file(self):
        if self.export_type == 'all':
            return self.__generate_for_restore()
        elif self.export_type == 'revise':
            return self.__generate_for_revise()
        elif self.export_type == 'wholesale':
            return self.__generate_for_wholesales()
        raise ValueError('Некорректный тип для выгрузки товара - %s' % self.export_type)


class InvoiceReportProcessor(object):
    def __init__(self, start_date, end_date):
        self.amount = 0
        self.overhead = 0
        self.start_date = start_date
        self.end_date = end_date
        self.invoices = []

    def __str__(self):
        return 'Список приемки товара с %s по %s' % (
            format_date(self.start_date), format_date(self.end_date))

    def process(self):
        self.invoices = Invoice.objects.select_related().prefetch_related('shipments').\
            filter(invoice_date__range=(self.start_date, self.end_date)).\
            order_by('-invoice_date')
        for invoice in self.invoices:
            self.amount += invoice.get_total_amount()
            self.overhead += invoice.overhead

        return self


class ReviseReportProcessor(object):
    def __init__(self, start_date, end_date):
        self.start_date = start_date
        self.end_date = end_date
        self.revises = []

    def process(self):
        self.revises = Revise.objects.select_related().prefetch_related('products_revise').\
            filter(status='ACCEPT').\
            filter(revise_date__range=(self.start_date, self.end_date)).\
            order_by('-revise_date')

        return self

    def __str__(self):
        return 'Список сверок товара за период с %s по %s' % (format_date(self.start_date), format_date(self.end_date))
