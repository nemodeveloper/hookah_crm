from datetime import datetime

from openpyxl import Workbook

from src.apps.cashbox.models import ProductSell
from src.base_components.excel.operation import BaseExcelOperation
from src.template_tags.common_tags import format_date, round_number


class ProductSellCheckOperation(BaseExcelOperation):

    def __init__(self, sell_id):
        super(ProductSellCheckOperation, self).__init__()
        self.sell = ProductSell.objects.get(id=sell_id)
        self.check_name = 'SellCheck_%s' % format_date(self.sell.sell_date, '%Y_%m_%d_%H_%M')

    def get_excel_check(self):

        book = Workbook()
        sheet = book.create_sheet(title='Чек по продаже', index=0)

        sheet.append(['Продажа от %s' % self.sell.get_verbose_sell_date()])
        sheet.append(['Список товаров продажи:'])
        sheet.append(['№', 'Группа', 'Категория', 'Вид', 'Наименование', 'Количество', 'Цена(шт)', 'Сумма'])

        cur = 1
        for shipment in self.sell.get_shipments():
            product = shipment.product
            product_category = product.product_kind.product_category
            sheet.append([cur, product_category.product_group.group_name,
                          product_category.category_name,
                          product.product_kind.kind_name,
                          product.product_name, shipment.product_count, shipment.cost_price, shipment.get_shipment_amount()])
            cur += 1
        sheet.append(['', '', '', '', '', '', 'Итого', self.sell.get_sell_amount()])
        sheet.append([])

        cur = 1
        sheet.append(['Оплата:'])
        sheet.append(['№', 'Тип', 'Сумма'])
        for payment in self.sell.get_payments():
            sheet.append([cur, payment.get_cash_type_display(), payment.cash])
            cur += 1
        sheet.append(['', 'Итого', self.sell.get_payment_amount()])

        self.post_process_sheet(sheet)

        return book


class SellCustomerReportExcelOperation(BaseExcelOperation):

    def __init__(self, report):
        super(SellCustomerReportExcelOperation, self).__init__()
        self.report = report
        self.file_name = 'SellCustomerReport_%s' % format_date(datetime.now(), '%Y_%m_%d_%H_%M')

    def get_excel_file(self):
        book = Workbook()
        sheet = book.create_sheet(title='Покупатели', index=0)

        def format_union_cell(start_row, end_row, column, value):
            sheet.merge_cells(start_row=start_row, end_row=end_row, start_column=column, end_column=column)
            sheet.cell(start_row, column).value = value
            self.centre_cell_value(sheet.cell(start_row, column))

        sheet.append(['', '', '', '', 'Отчет от %s по %s' % (format_date(self.report.start_date), format_date(self.report.end_date))])
        sheet.append(['№', 'Тип', 'Наименование', 'Категория', 'Вид', 'Товар', 'Кол-во(шт/гр/кг)', 'Цена(шт)', 'Всего(шт/гр/кг)', 'Сумма'])

        customer_number = 1
        start_row = 3
        current_row = start_row

        current_category_row = start_row
        current_kind_row = start_row

        for customer_name, customer_aggr in self.report.customers_aggr.items():
            customer_type_name = customer_aggr.customer.get_verbose_customer_type()

            customer_total_sum = customer_aggr.profit_report.total_cost_amount
            total_product_count = 0
            total_product_row = 0

            for group_name, group_aggr in customer_aggr.profit_report.groups_aggr.items():

                for category_aggr in group_aggr.categories_aggr:
                    category_name = category_aggr.category.category_name
                    category_product_count = 0

                    for kind_aggr in category_aggr.kinds_aggr:
                        kind_name = kind_aggr.kind.kind_name
                        total_product_count += kind_aggr.count
                        kind_product_count = 0

                        for product_aggr in kind_aggr.products_aggr:
                            product_name = product_aggr.product.product_name
                            product_count = product_aggr.count
                            product_cost = round_number(product_aggr.get_average_cost_price(), 2)

                            sheet.append(['', '', '', category_name, kind_name, product_name, product_count, product_cost, '', ''])
                            total_product_row += 1
                            category_product_count += 1
                            kind_product_count += 1

                        format_union_cell(current_kind_row, current_kind_row + kind_product_count - 1, 5, kind_name)
                        current_kind_row += kind_product_count

                    format_union_cell(current_category_row, current_category_row + category_product_count - 1, 4, category_name)
                    current_category_row += category_product_count

            if total_product_row > 0:
                customer_last_row = current_row + total_product_row - 1
            else:
                customer_last_row = current_row
                current_kind_row += 1
                current_category_row += 1
                sheet.append(['', '', '', '', '', '', '', '', '', ''])

            format_union_cell(current_row, customer_last_row, 1, customer_number)
            format_union_cell(current_row, customer_last_row, 2, customer_type_name)
            format_union_cell(current_row, customer_last_row, 3, customer_name)

            format_union_cell(current_row, customer_last_row, 9, round_number(total_product_count, 2))
            format_union_cell(current_row, customer_last_row, 10, round_number(customer_total_sum, 2))

            current_row = customer_last_row + 1

            customer_number += 1

        # Размер столбцов
        self.post_process_sheet(sheet)

        # Граница ячеек таблицы
        self.set_border(sheet, min_row=2, max_row=current_row - 1)

        # Центруем значения
        self.centre_value(sheet, min_row=2, max_row=current_row - 1)

        return book
